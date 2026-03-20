"""Full catalog import from Excel into PostgreSQL.

Reads the Excel catalog file (sterlitamak_services_catalog_v1.xlsx) and imports:
  - Professions (from direction_code/direction_name in DB_Import)
  - Service Groups (from group_code/group_name)
  - Service Subgroups (from subgroup_code/subgroup_name)
  - Shared Operations (from Shared_Ops sheet)
  - Service Items (all 330+ rows from DB_Import)
  - Coefficients (from Coeff_Template sheet)

Usage:
  python -m scripts.import_catalog path/to/catalog.xlsx
  python -m scripts.import_catalog path/to/catalog.xlsx --dry-run
  python -m scripts.import_catalog path/to/catalog.xlsx --drop-existing
"""

import argparse
import asyncio
import logging
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl не установлен: pip install openpyxl")
    sys.exit(1)

# Ensure UTF-8 on Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("import_catalog")


def read_sheet_rows(wb, sheet_name: str) -> list[dict]:
    """Read a sheet into list of dicts with header row as keys."""
    if sheet_name not in wb.sheetnames:
        logger.warning("Лист '%s' не найден, пропускаем", sheet_name)
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        result.append(d)
    return result


def safe_int(val, default=0) -> int:
    """Convert to int safely."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_str(val, default="") -> str:
    """Convert to string safely."""
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


def safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_bool(val, default=True) -> bool:
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "да", "y", "t")


async def run_import(excel_path: str, dry_run: bool = False, drop_existing: bool = False):
    """Main import logic."""
    from sqlalchemy import delete, select, text
    from app.database import get_async_session, get_engine
    from app.models.catalog import (
        Profession, ServiceGroup, ServiceItem, ServiceSubgroup, SharedOperation,
    )
    from app.models.coefficient import Coefficient

    path = Path(excel_path)
    if not path.exists():
        logger.error("Файл не найден: %s", path)
        sys.exit(1)

    logger.info("Открываем %s...", path.name)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    logger.info("Листы: %s", ", ".join(wb.sheetnames))

    # === Read sheets ===
    db_import_rows = read_sheet_rows(wb, "DB_Import")
    shared_ops_rows = read_sheet_rows(wb, "Shared_Ops")
    coeff_rows = read_sheet_rows(wb, "Coeff_Template")
    wb.close()

    if not db_import_rows:
        logger.error("Лист DB_Import пуст или не найден")
        sys.exit(1)

    logger.info("DB_Import: %d строк", len(db_import_rows))
    logger.info("Shared_Ops: %d строк", len(shared_ops_rows))
    logger.info("Coeff_Template: %d строк", len(coeff_rows))

    if dry_run:
        logger.info("DRY RUN — показываем первую строку DB_Import:")
        for k, v in db_import_rows[0].items():
            logger.info("  %s: %s", k, repr(v))
        logger.info("DRY RUN завершён, данные не записаны")
        return

    # === Database ===
    session_factory = get_async_session()

    async with session_factory() as session:
        if drop_existing:
            logger.warning("Удаляем существующие данные каталога...")
            await session.execute(delete(ServiceItem))
            await session.execute(delete(ServiceSubgroup))
            await session.execute(delete(ServiceGroup))
            await session.execute(delete(SharedOperation))
            await session.execute(delete(Coefficient))
            await session.execute(delete(Profession))
            await session.commit()
            logger.info("Существующие данные удалены")

        # === 1. Professions ===
        logger.info("--- Профессии ---")
        professions_map = {}  # code -> id
        seen_prof_codes = set()

        # Profession icons
        icon_map = {
            "EL": "⚡", "PL": "🔧", "FM": "🪑",
            "электрика": "⚡", "сантехника": "🔧", "сборка мебели": "🪑",
        }

        for row in db_import_rows:
            code = safe_str(row.get("direction_code"))
            name = safe_str(row.get("direction_name"))
            if not code or code in seen_prof_codes:
                continue
            seen_prof_codes.add(code)

            existing = (await session.execute(
                select(Profession).where(Profession.code == code)
            )).scalar_one_or_none()

            if existing:
                professions_map[code] = existing.id
                logger.info("  [существует] %s: %s (id=%d)", code, name, existing.id)
            else:
                prof = Profession(
                    code=code,
                    name=name,
                    icon=icon_map.get(code, icon_map.get(name.lower(), "🔧")),
                    sort_priority=len(professions_map),
                )
                session.add(prof)
                await session.flush()
                professions_map[code] = prof.id
                logger.info("  [создано] %s: %s (id=%d)", code, name, prof.id)

        logger.info("Профессий: %d", len(professions_map))

        # === 2. Service Groups ===
        logger.info("--- Группы ---")
        groups_map = {}  # code -> id
        seen_group_codes = set()

        for row in db_import_rows:
            code = safe_str(row.get("group_code"))
            name = safe_str(row.get("group_name"))
            prof_code = safe_str(row.get("direction_code"))
            if not code or code in seen_group_codes:
                continue
            seen_group_codes.add(code)

            prof_id = professions_map.get(prof_code)
            if not prof_id:
                logger.warning("  Профессия '%s' не найдена для группы '%s'", prof_code, code)
                continue

            existing = (await session.execute(
                select(ServiceGroup).where(ServiceGroup.code == code)
            )).scalar_one_or_none()

            if existing:
                groups_map[code] = existing.id
            else:
                grp = ServiceGroup(
                    profession_id=prof_id,
                    code=code,
                    name=name,
                    sort_priority=len(groups_map),
                )
                session.add(grp)
                await session.flush()
                groups_map[code] = grp.id

        logger.info("Групп: %d", len(groups_map))

        # === 3. Service Subgroups ===
        logger.info("--- Подгруппы ---")
        subgroups_map = {}  # code -> id
        seen_sub_codes = set()

        for row in db_import_rows:
            code = safe_str(row.get("subgroup_code"))
            name = safe_str(row.get("subgroup_name"))
            grp_code = safe_str(row.get("group_code"))
            if not code or code in seen_sub_codes:
                continue
            seen_sub_codes.add(code)

            grp_id = groups_map.get(grp_code)
            if not grp_id:
                continue

            existing = (await session.execute(
                select(ServiceSubgroup).where(ServiceSubgroup.code == code)
            )).scalar_one_or_none()

            if existing:
                subgroups_map[code] = existing.id
            else:
                sub = ServiceSubgroup(
                    group_id=grp_id,
                    code=code,
                    name=name,
                    sort_priority=len(subgroups_map),
                )
                session.add(sub)
                await session.flush()
                subgroups_map[code] = sub.id

        logger.info("Подгрупп: %d", len(subgroups_map))

        # === 4. Shared Operations ===
        logger.info("--- Общие операции ---")
        shared_count = 0

        for row in shared_ops_rows:
            code = safe_str(row.get("code") or row.get("op_code"))
            name = safe_str(row.get("name") or row.get("op_name"))
            if not code:
                continue

            existing = (await session.execute(
                select(SharedOperation).where(SharedOperation.code == code)
            )).scalar_one_or_none()

            if not existing:
                op = SharedOperation(
                    code=code,
                    name=name,
                    description=safe_str(row.get("description")),
                    typical_unit=safe_str(row.get("typical_unit") or row.get("unit")),
                    pricing_strategy=safe_str(row.get("pricing_strategy") or row.get("strategy")),
                )
                session.add(op)
                shared_count += 1

        if shared_count:
            await session.flush()
        logger.info("Общих операций добавлено: %d", shared_count)

        # === 5. Service Items (main catalog) ===
        logger.info("--- Позиции каталога ---")
        items_added = 0
        items_updated = 0
        items_skipped = 0

        for i, row in enumerate(db_import_rows):
            work_code = safe_str(row.get("work_code") or row.get("code"))
            if not work_code:
                items_skipped += 1
                continue

            # Resolve foreign keys
            prof_code = safe_str(row.get("direction_code"))
            grp_code = safe_str(row.get("group_code"))
            sub_code = safe_str(row.get("subgroup_code"))

            prof_id = professions_map.get(prof_code)
            grp_id = groups_map.get(grp_code)
            sub_id = subgroups_map.get(sub_code) if sub_code else None

            if not prof_id or not grp_id:
                items_skipped += 1
                continue

            # Build search_text from multiple fields
            name = safe_str(row.get("work_name") or row.get("name"))
            aliases = safe_str(row.get("aliases"))
            hashtags = safe_str(row.get("hashtags"))
            raw_search = safe_str(row.get("search_text"))

            search_text = " ".join(filter(None, [
                name.lower(),
                aliases.lower() if aliases else "",
                hashtags.replace("#", " ").lower() if hashtags else "",
                raw_search.lower() if raw_search else "",
            ])).strip()

            # Check existing
            existing = (await session.execute(
                select(ServiceItem).where(ServiceItem.code == work_code)
            )).scalar_one_or_none()

            item_data = dict(
                sort_order=safe_int(row.get("sort_order"), i),
                profession_id=prof_id,
                group_id=grp_id,
                subgroup_id=sub_id,
                code=work_code,
                slug=safe_str(row.get("slug") or work_code.lower().replace("-", "_")),
                name=name,
                unit=safe_str(row.get("unit"), "шт"),
                price_min=safe_int(row.get("price_min_rub") or row.get("price_min")),
                price_max=safe_int(row.get("price_max_rub") or row.get("price_max")),
                price_recommended=safe_int(row.get("price_rec_rub") or row.get("price_recommended")),
                currency=safe_str(row.get("currency"), "RUB"),
                record_type=safe_str(row.get("record_type"), "atomic"),
                calc_strategy=safe_str(row.get("calc_strategy"), "PER_UNIT"),
                selection_mode=safe_str(row.get("selection_mode"), "quantity"),
                complexity=safe_str(row.get("complexity")) or None,
                confidence=safe_str(row.get("confidence")) or None,
                labor_only=safe_bool(row.get("labor_only"), True),
                aliases=aliases or None,
                hashtags=hashtags or None,
                search_text=search_text or None,
                shared_ops=safe_str(row.get("shared_ops")) or None,
                excludes=safe_str(row.get("excludes")) or None,
                estimator_fields=safe_str(row.get("estimator_fields")) or None,
                note=safe_str(row.get("note")) or None,
                source_1=safe_str(row.get("source_1")) or None,
                source_2=safe_str(row.get("source_2")) or None,
                city=safe_str(row.get("city")) or None,
                region=safe_str(row.get("region")) or None,
                price_updated_at=safe_str(row.get("price_updated_at")) or None,
                is_active=safe_bool(row.get("active"), True),
            )

            if existing:
                for k, v in item_data.items():
                    setattr(existing, k, v)
                items_updated += 1
            else:
                session.add(ServiceItem(**item_data))
                items_added += 1

            # Flush every 50 items to avoid large batches
            if (i + 1) % 50 == 0:
                await session.flush()
                logger.info("  ... обработано %d / %d", i + 1, len(db_import_rows))

        await session.flush()
        logger.info("Позиций: добавлено=%d, обновлено=%d, пропущено=%d",
                     items_added, items_updated, items_skipped)

        # === 6. Coefficients ===
        logger.info("--- Коэффициенты ---")
        coeff_count = 0

        for row in coeff_rows:
            key = safe_str(row.get("coef_key") or row.get("key"))
            if not key:
                continue

            existing = (await session.execute(
                select(Coefficient).where(Coefficient.coef_key == key)
            )).scalar_one_or_none()

            if not existing:
                coef = Coefficient(
                    coef_type=safe_str(row.get("coef_type") or row.get("type"), "other"),
                    coef_key=key,
                    label=safe_str(row.get("label") or row.get("name"), key),
                    multiplier=safe_float(row.get("multiplier") or row.get("value"), 1.0),
                    applies_to=safe_str(row.get("applies_to")) or None,
                    when_use=safe_str(row.get("when_use") or row.get("description")) or None,
                    note=safe_str(row.get("note")) or None,
                    sort_priority=safe_int(row.get("sort_priority") or row.get("sort"), 0),
                )
                session.add(coef)
                coeff_count += 1

        if coeff_count:
            await session.flush()
        logger.info("Коэффициентов добавлено: %d", coeff_count)

        # === Commit ===
        await session.commit()
        logger.info("=== Импорт завершён успешно ===")

        # Summary
        total_items = (await session.execute(
            select(sa_func.count(ServiceItem.id))
        )).scalar() if False else items_added + items_updated

        logger.info("ИТОГО:")
        logger.info("  Профессий:       %d", len(professions_map))
        logger.info("  Групп:           %d", len(groups_map))
        logger.info("  Подгрупп:        %d", len(subgroups_map))
        logger.info("  Общих операций:  %d", shared_count)
        logger.info("  Позиций:         %d (новых: %d, обновлённых: %d)",
                     items_added + items_updated, items_added, items_updated)
        logger.info("  Коэффициентов:   %d", coeff_count)


def main():
    parser = argparse.ArgumentParser(description="Импорт каталога из Excel в БД МастерБот")
    parser.add_argument("excel_path", help="Путь к Excel-файлу каталога")
    parser.add_argument("--dry-run", action="store_true", help="Только показать данные, не записывать")
    parser.add_argument("--drop-existing", action="store_true", help="Удалить существующие данные перед импортом")
    args = parser.parse_args()

    asyncio.run(run_import(args.excel_path, dry_run=args.dry_run, drop_existing=args.drop_existing))


if __name__ == "__main__":
    main()
