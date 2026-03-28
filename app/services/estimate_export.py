"""Estimate export to PDF and XLSX with QR code for payment.

Generates professional-looking Russian-standard documents with:
- Master's personal data header
- Line items table with totals
- Discount breakdown
- QR code (ST-00012 banking standard) from master's bank details
- Footer with payment instructions and contacts
"""

import io
import logging
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path

logger = logging.getLogger(__name__)

QR_FORMAT_PREFIX = "ST00011"
QR_CHARSET = "cp1251"
QR_REQUIRED_BANK_FIELDS: tuple[tuple[str, str], ...] = (
    ("payment_recipient", "Получатель"),
    ("settlement_account", "Расчетный счет"),
    ("bank_name", "Банк"),
    ("bik", "БИК"),
    ("correspondent_account", "Корреспондентский счет"),
)

PDF_FONT_CANDIDATES: tuple[tuple[str, str, str], ...] = (
    (
        "DejaVuSans",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ),
    (
        "Arial",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ),
    (
        "Tahoma",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/tahomabd.ttf",
    ),
)


@dataclass
class ExportProfile:
    """Master's profile data for document generation."""
    full_name: str = ""
    phone: str = ""
    email: str = ""
    telegram_username: str = ""
    company_name: str = ""
    inn: str = ""
    address: str = ""
    specialization: str = ""
    # Bank details
    bank_name: str = ""
    bik: str = ""
    correspondent_account: str = ""
    settlement_account: str = ""
    card_number: str = ""
    sbp_phone: str = ""
    payment_recipient: str = ""


@dataclass
class ExportLineItem:
    number: int
    name: str
    unit: str
    quantity: float
    unit_price: int
    coefficients: str  # human-readable
    subtotal: int


@dataclass
class ExportEstimate:
    estimate_id: int
    version: int
    status: str
    created_at: str
    items: list[ExportLineItem]
    total: int
    discount: int
    final: int
    discount_reason: str = ""
    note: str = ""
    client_name: str = ""


def _clean_qr_value(value: str | None, *, digits_only: bool = False, max_len: int | None = None) -> str:
    if not value:
        return ""
    cleaned = " ".join(str(value).replace("\r", " ").replace("\n", " ").split())
    cleaned = cleaned.replace("|", "/").replace("=", "-")
    if digits_only:
        cleaned = "".join(ch for ch in cleaned if ch.isdigit())
    if max_len is not None:
        cleaned = cleaned[:max_len]
    return cleaned


def get_missing_bank_qr_fields(profile: ExportProfile) -> list[str]:
    missing: list[str] = []
    for field, label in QR_REQUIRED_BANK_FIELDS:
        if not _clean_qr_value(getattr(profile, field, "")):
            missing.append(label)
    return missing


def has_bank_qr_details(profile: ExportProfile) -> bool:
    return not get_missing_bank_qr_fields(profile)


def _build_qr_payload(
    profile: ExportProfile,
    amount: int | None = None,
    purpose: str | None = None,
) -> str:
    """Build a bank QR payload from full requisites."""
    if not has_bank_qr_details(profile):
        return ""

    recipient = _clean_qr_value(profile.payment_recipient or profile.full_name, max_len=160)
    settlement_account = _clean_qr_value(profile.settlement_account, digits_only=True, max_len=20)
    bank_name = _clean_qr_value(profile.bank_name, max_len=160)
    bik = _clean_qr_value(profile.bik, digits_only=True, max_len=9)
    correspondent_account = _clean_qr_value(profile.correspondent_account, digits_only=True, max_len=20)

    parts = [
        QR_FORMAT_PREFIX,
        f"Name={recipient}",
        f"PersonalAcc={settlement_account}",
        f"BankName={bank_name}",
        f"BIC={bik}",
        f"CorrespAcc={correspondent_account}",
    ]
    inn = _clean_qr_value(profile.inn, digits_only=True, max_len=12)
    if inn:
        parts.append(f"PayeeINN={inn}")
    clean_purpose = _clean_qr_value(purpose, max_len=210)
    if clean_purpose:
        parts.append(f"Purpose={clean_purpose}")
    if amount is not None:
        parts.append(f"Sum={int(amount) * 100}")
    return "|".join(parts)


def _build_sbp_qr_payload(profile: ExportProfile, amount: int, estimate_id: int) -> str:
    """Build simplified QR for SBP (СБП) phone transfer."""
    phone = profile.sbp_phone or profile.phone
    if not phone:
        return ""
    # Simplified: just put phone + amount in text for SBP-compatible apps
    return f"ST00012|Name={profile.payment_recipient or profile.full_name}|PersonalAcc={phone}|Sum={amount}00|Purpose=Смета #{estimate_id}"


def _generate_qr_image(data: str, size: int = 180) -> bytes | None:
    """Generate QR code as PNG bytes."""
    if not data:
        return None
    try:
        import qrcode
        from qrcode.image.pil import PilImage

        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=max(4, size // 48),
            border=4,
        )
        qr.add_data(data.encode(QR_CHARSET, errors="replace"))
        qr.make(fit=True)
        img = qr.make_image(image_factory=PilImage, fill_color="black", back_color="white")

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception as e:
        logger.warning("QR generation failed: %s", e)
        return None


def _money(amount: int) -> str:
    """Format money Russian-style: 12500 → '12 500'."""
    return f"{amount:,}".replace(",", " ")


@lru_cache(maxsize=1)
def _get_pdf_font_names() -> tuple[str, str]:
    """Resolve a Unicode TTF family so Cyrillic renders correctly in PDF."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    registered = set(pdfmetrics.getRegisteredFontNames())
    for family_name, regular_path_raw, bold_path_raw in PDF_FONT_CANDIDATES:
        regular_path = Path(regular_path_raw)
        bold_path = Path(bold_path_raw)
        if not regular_path.exists() or not bold_path.exists():
            continue

        regular_name = f"{family_name}-UTF8"
        bold_name = f"{family_name}-UTF8-Bold"
        if regular_name not in registered:
            pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
            registered.add(regular_name)
        if bold_name not in registered:
            pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
            registered.add(bold_name)
        return regular_name, bold_name

    logger.warning("No Unicode TTF font found for PDF export; falling back to Helvetica")
    return "Helvetica", "Helvetica-Bold"


# ─── PDF Export ──────────────────────────────────────────────────────────────

def export_pdf(estimate: ExportEstimate, profile: ExportProfile) -> bytes:
    """Generate professional PDF estimate document."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buf = io.BytesIO()
    font_regular, font_bold = _get_pdf_font_names()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    style_title = ParagraphStyle(
        "DocTitle", parent=styles["Heading1"],
        fontName=font_bold,
        fontSize=16, spaceAfter=4 * mm, alignment=1,  # center
    )
    style_subtitle = ParagraphStyle(
        "DocSubtitle", parent=styles["Normal"],
        fontName=font_regular,
        fontSize=9, textColor=colors.grey, alignment=1, spaceAfter=6 * mm,
    )
    style_header = ParagraphStyle(
        "Header", parent=styles["Normal"],
        fontName=font_regular,
        fontSize=9, leading=13, textColor=colors.HexColor("#333333"),
    )
    style_footer = ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontName=font_regular,
        fontSize=8, textColor=colors.grey, leading=11,
    )
    style_note = ParagraphStyle(
        "Note", parent=styles["Normal"],
        fontName=font_regular,
        fontSize=8, textColor=colors.HexColor("#666666"),
        leading=11, spaceBefore=3 * mm,
    )

    elements = []

    # ─── Title ───
    elements.append(Paragraph("СМЕТА НА ВЫПОЛНЕНИЕ РАБОТ", style_title))
    elements.append(Paragraph(
        f"№ {estimate.estimate_id} (версия {estimate.version}) от {estimate.created_at}",
        style_subtitle,
    ))

    # ─── Master info block ───
    master_lines = []
    if profile.company_name:
        master_lines.append(f"<b>Исполнитель:</b> {profile.company_name}")
    if profile.full_name:
        master_lines.append(f"<b>Мастер:</b> {profile.full_name}")
    if profile.specialization:
        master_lines.append(f"<b>Специализация:</b> {profile.specialization}")
    if profile.phone:
        master_lines.append(f"<b>Тел.:</b> {profile.phone}")
    if profile.email:
        master_lines.append(f"<b>Email:</b> {profile.email}")
    if profile.telegram_username:
        master_lines.append(f"<b>Telegram:</b> @{profile.telegram_username.lstrip('@')}")
    if profile.inn:
        master_lines.append(f"<b>ИНН:</b> {profile.inn}")
    if profile.address:
        master_lines.append(f"<b>Адрес:</b> {profile.address}")

    if master_lines:
        elements.append(Paragraph("<br/>".join(master_lines), style_header))
        elements.append(Spacer(1, 3 * mm))

    # Client info
    if estimate.client_name:
        elements.append(Paragraph(
            f"<b>Заказчик:</b> {estimate.client_name}", style_header,
        ))
        elements.append(Spacer(1, 3 * mm))

    # ─── Line items table ───
    header_row = ["№", "Наименование работ", "Ед.", "Кол-во", "Цена, ₽", "Коэфф.", "Сумма, ₽"]
    table_data = [header_row]

    for item in estimate.items:
        table_data.append([
            str(item.number),
            item.name,
            item.unit,
            f"{item.quantity:g}",
            _money(item.unit_price),
            item.coefficients or "—",
            _money(item.subtotal),
        ])

    # Column widths (total ~180mm for A4 with margins)
    col_widths = [8 * mm, 62 * mm, 14 * mm, 14 * mm, 22 * mm, 22 * mm, 24 * mm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B5797")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        # Body
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, -1), font_regular),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),  # №
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),  # unit+
        ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),  # total
        ("ALIGN", (-3, 1), (-3, -1), "RIGHT"),  # price
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#2B5797")),
        # Alternate rows
        *[
            ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F5F7FA"))
            for i in range(2, len(table_data), 2)
        ],
    ]))
    elements.append(table)
    elements.append(Spacer(1, 4 * mm))

    # ─── Totals ───
    totals_data = []
    totals_data.append(["", "Итого:", _money(estimate.total) + " ₽"])
    if estimate.discount > 0:
        discount_text = f"Скидка ({estimate.discount_reason}):" if estimate.discount_reason else "Скидка:"
        totals_data.append(["", discount_text, f"−{_money(estimate.discount)} ₽"])
    totals_data.append(["", "К ОПЛАТЕ:", _money(estimate.final) + " ₽"])

    totals_table = Table(totals_data, colWidths=[100 * mm, 40 * mm, 30 * mm])
    totals_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), font_regular),
        ("FONTNAME", (1, 0), (1, -1), font_bold),
        ("FONTNAME", (2, 0), (2, -1), font_bold),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        # Last row bold and colored
        ("FONTSIZE", (1, -1), (-1, -1), 11),
        ("TEXTCOLOR", (1, -1), (-1, -1), colors.HexColor("#2B5797")),
        ("LINEABOVE", (1, -1), (-1, -1), 1, colors.HexColor("#2B5797")),
        ("TOPPADDING", (0, -1), (-1, -1), 4),
    ]))
    elements.append(totals_table)

    # ─── QR code + payment details ───
    qr_payload = _build_qr_payload(
        profile,
        amount=estimate.final,
        purpose=f"Оплата по смете #{estimate.estimate_id}",
    )
    qr_image_bytes = _generate_qr_image(qr_payload)
    has_bank_details = bool(profile.bank_name or profile.sbp_phone or profile.card_number)

    if has_bank_details:
        elements.append(Spacer(1, 6 * mm))
        elements.append(Paragraph(
            "<b>РЕКВИЗИТЫ ДЛЯ ОПЛАТЫ</b>",
            ParagraphStyle("PayTitle", parent=styles["Normal"], fontName=font_bold, fontSize=10,
                           textColor=colors.HexColor("#2B5797"), spaceBefore=2 * mm),
        ))
        elements.append(Spacer(1, 2 * mm))

        # Build payment info text
        pay_lines = []
        if profile.payment_recipient:
            pay_lines.append(f"<b>Получатель:</b> {profile.payment_recipient}")
        if profile.bank_name:
            pay_lines.append(f"<b>Банк:</b> {profile.bank_name}")
        if profile.settlement_account:
            pay_lines.append(f"<b>Р/с:</b> {profile.settlement_account}")
        if profile.correspondent_account:
            pay_lines.append(f"<b>Корр. счёт:</b> {profile.correspondent_account}")
        if profile.bik:
            pay_lines.append(f"<b>БИК:</b> {profile.bik}")
        if profile.inn:
            pay_lines.append(f"<b>ИНН:</b> {profile.inn}")
        if profile.card_number:
            pay_lines.append(f"<b>Карта:</b> {profile.card_number}")
        if profile.sbp_phone:
            pay_lines.append(f"<b>СБП (телефон):</b> {profile.sbp_phone}")

        pay_text = Paragraph("<br/>".join(pay_lines), style_header)

        if qr_image_bytes:
            # Side-by-side: QR on left, details on right
            qr_img = Image(io.BytesIO(qr_image_bytes), width=40 * mm, height=40 * mm)
            qr_table = Table(
                [[qr_img, pay_text]],
                colWidths=[45 * mm, 120 * mm],
            )
            qr_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 4 * mm),
            ]))
            elements.append(qr_table)
        else:
            elements.append(pay_text)

    # ─── Notes ───
    if estimate.note:
        elements.append(Paragraph(f"<b>Примечание:</b> {estimate.note}", style_note))

    # ─── Footer / legal ───
    elements.append(Spacer(1, 8 * mm))
    footer_lines = [
        "* Цены указаны в российских рублях (₽). Стоимость может быть скорректирована при изменении объёма работ.",
        "* Срок действия сметы — 30 дней с даты составления.",
        "* Гарантия на выполненные работы — по договорённости с мастером.",
    ]
    if profile.sbp_phone:
        footer_lines.append(
            f"* Оплата через СБП: переведите {_money(estimate.final)} ₽ на номер {profile.sbp_phone} ({profile.bank_name or 'банк мастера'})."
        )
    for line in footer_lines:
        elements.append(Paragraph(line, style_footer))

    # Build
    doc.build(elements)
    return buf.getvalue()


# ─── XLSX Export ─────────────────────────────────────────────────────────────

def export_xlsx(estimate: ExportEstimate, profile: ExportProfile) -> bytes:
    """Generate professional XLSX estimate document."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Смета #{estimate.estimate_id}"

    # Styles
    header_font = Font(name="Arial", size=14, bold=True, color="2B5797")
    subtitle_font = Font(name="Arial", size=9, color="888888")
    label_font = Font(name="Arial", size=9, bold=True)
    value_font = Font(name="Arial", size=9)
    th_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    th_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    total_font = Font(name="Arial", size=11, bold=True, color="2B5797")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Column widths
    widths = {"A": 5, "B": 45, "C": 8, "D": 8, "E": 14, "F": 12, "G": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ─── Title ───
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "СМЕТА НА ВЫПОЛНЕНИЕ РАБОТ"
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = f"№ {estimate.estimate_id} (версия {estimate.version}) от {estimate.created_at}"
    ws[f"A{row}"].font = subtitle_font
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 2

    # ─── Master info ───
    def add_field(label, value):
        nonlocal row
        if value:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = label_font
            ws.merge_cells(f"A{row}:B{row}")
            ws[f"C{row}"] = value
            ws[f"C{row}"].font = value_font
            ws.merge_cells(f"C{row}:G{row}")
            row += 1

    if profile.company_name:
        add_field("Исполнитель:", profile.company_name)
    add_field("Мастер:", profile.full_name)
    add_field("Специализация:", profile.specialization)
    add_field("Тел.:", profile.phone)
    add_field("Email:", profile.email)
    if profile.telegram_username:
        add_field("Telegram:", f"@{profile.telegram_username.lstrip('@')}")
    add_field("ИНН:", profile.inn)
    add_field("Адрес:", profile.address)

    if estimate.client_name:
        row += 1
        add_field("Заказчик:", estimate.client_name)

    row += 1

    # ─── Table header ───
    headers = ["№", "Наименование работ", "Ед.", "Кол-во", "Цена, ₽", "Коэфф.", "Сумма, ₽"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = th_font
        cell.fill = th_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    row += 1

    # ─── Line items ───
    for i, item in enumerate(estimate.items):
        values = [
            item.number, item.name, item.unit,
            item.quantity, item.unit_price,
            item.coefficients or "—", item.subtotal,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = value_font
            cell.border = thin_border
            if col_idx in (1, 3, 4, 6):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (5, 7):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            if i % 2 == 1:
                cell.fill = alt_fill
        row += 1

    row += 1

    # ─── Totals ───
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"F{row}"] = "Итого:"
    ws[f"F{row}"].font = label_font
    ws[f"F{row}"].alignment = Alignment(horizontal="right")
    ws[f"G{row}"] = estimate.total
    ws[f"G{row}"].font = label_font
    ws[f"G{row}"].number_format = '#,##0'
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    row += 1

    if estimate.discount > 0:
        ws[f"F{row}"] = "Скидка:"
        ws[f"F{row}"].font = value_font
        ws[f"F{row}"].alignment = Alignment(horizontal="right")
        ws[f"G{row}"] = -estimate.discount
        ws[f"G{row}"].font = value_font
        ws[f"G{row}"].number_format = '#,##0'
        ws[f"G{row}"].alignment = Alignment(horizontal="right")
        row += 1

    ws[f"F{row}"] = "К ОПЛАТЕ:"
    ws[f"F{row}"].font = total_font
    ws[f"F{row}"].alignment = Alignment(horizontal="right")
    ws[f"G{row}"] = estimate.final
    ws[f"G{row}"].font = total_font
    ws[f"G{row}"].number_format = '#,##0'
    ws[f"G{row}"].alignment = Alignment(horizontal="right")
    row += 2

    # ─── Bank details ───
    has_bank = bool(profile.bank_name or profile.sbp_phone or profile.card_number)
    if has_bank:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = "РЕКВИЗИТЫ ДЛЯ ОПЛАТЫ"
        ws[f"A{row}"].font = Font(name="Arial", size=10, bold=True, color="2B5797")
        row += 1
        add_field("Получатель:", profile.payment_recipient)
        add_field("Банк:", profile.bank_name)
        add_field("Р/с:", profile.settlement_account)
        add_field("Корр. счёт:", profile.correspondent_account)
        add_field("БИК:", profile.bik)
        add_field("ИНН:", profile.inn)
        add_field("Карта:", profile.card_number)
        add_field("СБП (телефон):", profile.sbp_phone)
        row += 1

    # ─── QR code image ───
    qr_payload = _build_qr_payload(
        profile,
        amount=estimate.final,
        purpose=f"Оплата по смете #{estimate.estimate_id}",
    )
    qr_bytes = _generate_qr_image(qr_payload, size=200)
    if qr_bytes and has_bank:
        from openpyxl.drawing.image import Image as XlImage
        qr_img = XlImage(io.BytesIO(qr_bytes))
        qr_img.width = 150
        qr_img.height = 150
        ws.add_image(qr_img, f"A{row}")
        row += 9  # ~9 rows for the image height

    # ─── Notes ───
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "* Цены указаны в российских рублях (₽). Стоимость может быть скорректирована при изменении объёма работ."
    ws[f"A{row}"].font = Font(name="Arial", size=7, color="999999")
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "* Срок действия сметы — 30 дней с даты составления."
    ws[f"A{row}"].font = Font(name="Arial", size=7, color="999999")
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "* Гарантия на выполненные работы — по договорённости с мастером."
    ws[f"A{row}"].font = Font(name="Arial", size=7, color="999999")

    if estimate.note:
        row += 1
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = f"Примечание: {estimate.note}"
        ws[f"A{row}"].font = Font(name="Arial", size=8, italic=True, color="666666")

    # Print setup
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── QR code generation ──────────────────────────────────────────────────────

def generate_payment_qr(
    profile: ExportProfile,
    amount: int | None = None,
    estimate_id: int | None = None,
    *,
    purpose: str | None = None,
) -> dict:
    """Generate payment QR code and return all payment info."""
    qr_payload = _build_qr_payload(
        profile,
        amount=amount,
        purpose=purpose or (f"Оплата по смете #{estimate_id}" if estimate_id is not None else None),
    )
    qr_bytes = _generate_qr_image(qr_payload)

    import base64
    qr_base64 = base64.b64encode(qr_bytes).decode() if qr_bytes else None

    return {
        "qr_data": qr_payload,
        "qr_image": qr_base64,
        "amount": amount,
        "recipient": profile.payment_recipient or profile.full_name,
        "bank": profile.bank_name,
        "account": profile.settlement_account,
        "bik": profile.bik,
        "correspondent_account": profile.correspondent_account,
        "card": profile.card_number,
        "sbp_phone": profile.sbp_phone,
        "inn": profile.inn,
        "has_bank_qr": bool(qr_payload and qr_base64),
        "missing_bank_fields": get_missing_bank_qr_fields(profile),
    }
